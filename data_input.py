import pandas as pd
import numpy as np
from openpyxl import load_workbook


class data_input:
    def __init__(self, file):

        # get all file date on init
        self.input_dict = self.get_all_tables(file)

    def get_all_tables(self, filename):
        """ Get all tables from a given workbook. Returns a dictionary of tables.
            Requires a filename, which includes the file path and filename.
            https://stackoverflow.com/questions/43941365/openpyxl-read-tables-from-existing-data-book-example
            """

        # Load the workbook, from the filename
        wb = load_workbook(filename=filename, read_only=False, keep_vba=False, data_only=True, keep_links=False)

        # Initialize the dictionary of tables
        tables_dict = {}
        data_dict = {}

        # Go through each worksheet in the workbook
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]

            # Get each table in the worksheet
            for tbl in ws.tables.values():
                # First, add some info about the table to the dictionary
                tables_dict[tbl.name] = {
                    'table_name': tbl.name,
                    'worksheet': ws_name,
                    'num_cols': len(tbl.tableColumns),
                    'table_range': tbl.ref}

                # Grab the 'data' from the table
                data = ws[tbl.ref]

                # Create 2D List containing table data
                rows_list = []
                for row in data:
                    # Get a list of all columns in each row
                    cols = []
                    for col in row:
                        cols.append(col.value)
                    rows_list.append(cols)

                # Make standard data table available as dict #
                # detect standard input table by header names
                if rows_list[0][1] == "Unit" and rows_list[0][2] == "Value":

                    data_dict = {}  # create empty dict to represent dataframe

                    for data_row in rows_list[1:]:  # skip header row

                        if len(rows_list[0]) >= 5:  # Check if variation is given
                            if data_row[3] is not None and rows_list[0][3] == "Variation" and rows_list[0][4] == "Step":

                                variations = []
                                i = 1 - data_row[3]  # Set counter to lower bound

                                while i < 1 + data_row[3]:  # While counter is smaller than upper bound
                                    variations.append(i)  # Add new variation value
                                    # Increment counter by step
                                    i = i + data_row[4]

                                # Add upper bound to end of variations
                                variations.append(1 + data_row[3])

                                data_dict.update(
                                    # add name as key, value and variation as list elements
                                    {data_row[0]: {"Value": data_row[2],
                                                   "Variations": variations
                                                   }}
                                )
                            else:  # If no Variation given, add only value
                                data_dict.update(
                                    {data_row[0]: data_row[2]}  # add name and value to the dict
                                )

                        else:  # If no Variation given, add only value
                            data_dict.update(
                                {data_row[0]: data_row[2]}  # add name and value to the dict
                            )

                    # Add dict to dictionary of tables
                    tables_dict[tbl.name]['dict'] = data_dict

                else:  # if not standard format, save as dataframe

                    # Create a pandas dataframe from the rows_list.
                    # The first row is the column names
                    df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

                    # Add the dataframe to the dictionary of tables
                    tables_dict[tbl.name]['dataframe'] = df

        return tables_dict
